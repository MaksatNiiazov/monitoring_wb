from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
import re
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from monitoring.models import (
    DailyCampaignProductStat,
    DailyProductKeywordStat,
    DailyProductMetrics,
    DailyProductNote,
    DailyProductStock,
    DailyWarehouseStock,
    Product,
    ProductKeyword,
    ProductEconomicsVersion,
)
from monitoring.services.config import get_monitoring_settings
from monitoring.services.exporters import exporter_rows
from monitoring.services.reports import (
    MetricCell,
    ResolvedEconomics,
    build_dashboard_context,
    build_product_report,
    decimalize,
    has_metric_cell_data,
    normalize_search_text,
    normalize_warehouse_name,
)

BASE_BLOCK_HEIGHT = 50
BASE_KEYWORD_ROWS = 0
BLOCK_WIDTH = 9
BLOCK_GAP = 0
SHEETS_MAX_TITLE_LENGTH = 100


@dataclass
class MonitoringSheetPayload:
    title: str
    rows: list[list[Any]]
    kind: str = "product"
    product_id: int | None = None
    block_dates: list[date] | None = None


THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BLOCK_EDGE_SIDE = Side(style="thick", color="000000")
BLOCK_LEFT_BORDER = Border(left=BLOCK_EDGE_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BLOCK_RIGHT_BORDER = Border(left=THIN_SIDE, right=BLOCK_EDGE_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
NO_BORDER = Border()
HEADER_FILL = PatternFill("solid", fgColor="15324D")
SUBHEADER_FILL = PatternFill("solid", fgColor="E7EEF8")
SECTION_FILL = PatternFill("solid", fgColor="EEF3F7")
FORMULA_FILL = PatternFill("solid", fgColor="F1EAFE")
META_FILL = PatternFill("solid", fgColor="F7F9FC")
BLOCK_SEPARATOR_FILL = PatternFill("solid", fgColor="DDE5F2")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(color="15324D", bold=True)


def normalize_title(value: str) -> str:
    normalized = re.sub(r"[\\/?*\[\]:]", "_", value).strip()
    return normalized[:SHEETS_MAX_TITLE_LENGTH] or "Sheet1"


def monitoring_sheet_title(product: Product) -> str:
    base = product.vendor_code or product.title or f"WB {product.nm_id}"
    return normalize_title(base)


def _money(value: Decimal | int | float | str | None, optional: bool = False) -> float | str:
    """Форматирование денежного значения. optional=True возвращает '' для нуля."""
    number = decimalize(value)
    if optional and number == 0:
        return ""
    return float(number.quantize(Decimal("0.01")))


def _int(value: Decimal | int | float | str | None, optional: bool = False) -> int | str:
    """Форматирование целого. optional=True возвращает '' для нуля."""
    number = decimalize(value)
    if optional and number == 0:
        return ""
    return int(number)


def _fraction(value: Decimal | int | float | str | None, optional: bool = False) -> float | str:
    """Форматирование доли (0-1 → 0-100%). optional=True возвращает '' для нуля."""
    number = decimalize(value)
    # Нормализуем: если значение между -1 и 1 (не включая 0), умножаем на 100
    if Decimal("-1") < number < Decimal("1") and number != 0:
        number = number * Decimal("100")
    if optional and number == 0:
        return ""
    return float((number / Decimal("100")).quantize(Decimal("0.0001")))


def _bool(value: bool) -> str:
    return "Да" if value else "Нет"


def _keyword_offset(keyword_rows: list[dict[str, Any]]) -> int:
    return len(keyword_rows or [])


def _block_height(keyword_rows: list[dict[str, Any]]) -> int:
    return BASE_BLOCK_HEIGHT + _keyword_offset(keyword_rows)


def _normalize_keyword_rows_count(reports: list[dict[str, Any]]) -> None:
    if not reports:
        return
    keyword_texts: list[str] = []
    seen_keywords: set[str] = set()
    target = 0
    for report in reports:
        keyword_rows = report.get("keyword_rows") or []
        target = max(target, len(keyword_rows))
        for keyword_row in keyword_rows:
            query_text = str(keyword_row.get("query_text") or "").strip()
            normalized_query = normalize_search_text(query_text)
            if not normalized_query or normalized_query in seen_keywords:
                continue
            seen_keywords.add(normalized_query)
            keyword_texts.append(query_text)
    target = max(target, len(keyword_texts))

    for report in reports:
        keyword_rows = report.get("keyword_rows") or []
        rows_by_query: dict[str, dict[str, Any]] = {}
        for keyword_row in keyword_rows:
            normalized_query = normalize_search_text(str(keyword_row.get("query_text") or ""))
            if normalized_query:
                rows_by_query.setdefault(normalized_query, keyword_row)

        aligned_rows: list[dict[str, Any]] = []
        for query_text in keyword_texts:
            normalized_query = normalize_search_text(query_text)
            source_row = rows_by_query.get(normalized_query)
            if source_row:
                aligned_rows.append({**source_row, "query_text": query_text})
            else:
                aligned_rows.append(
                    {
                        "query_text": query_text,
                        "has_data": False,
                        "frequency": None,
                        "organic_position": None,
                        "boosted_position": None,
                        "boosted_ctr": None,
                    }
                )

        missing = max(0, target - len(aligned_rows))
        for _ in range(missing):
            aligned_rows.append(
                {
                    "query_text": "",
                    "has_data": False,
                    "frequency": None,
                    "organic_position": None,
                    "boosted_position": None,
                    "boosted_ctr": None,
                }
            )
        report["keyword_rows"] = aligned_rows


def _cell_ref(*, start_row: int, start_col: int, relative_row: int, relative_col: int) -> str:
    return f"{get_column_letter(start_col + relative_col - 1)}{start_row + relative_row - 1}"


def _cell_formula(
    *,
    start_row: int,
    start_col: int,
    template: str,
    refs: dict[str, tuple[int, int]],
) -> str:
    resolved = {
        key: _cell_ref(start_row=start_row, start_col=start_col, relative_row=rel_row, relative_col=rel_col)
        for key, (rel_row, rel_col) in refs.items()
    }
    return "=" + template.format(**resolved)


def _metric_cell(report: dict[str, Any], group: str, zone: str) -> MetricCell:
    return report["cells"].get((group, zone), MetricCell())


def _block_header(report: dict[str, Any]) -> str:
    stock_date = report["stock_date"]
    return f"{stock_date:%d.%m.%Y}"


def _spp_delta_parts(
    report: dict[str, Any],
    previous_report: dict[str, Any] | None,
) -> tuple[str, float | str]:
    current = decimalize(report["note"].spp_percent)
    if current == 0:
        return ("", "")
    if previous_report is None:
        return ("Без изменений", "")
    previous = decimalize(previous_report["note"].spp_percent)
    if previous == 0:
        return ("Без изменений", "")
    delta = current - previous
    if delta == 0:
        return ("Без изменений", "")
    label = "Вырос на" if delta > 0 else "Упал на"
    return (label, _fraction(delta.copy_abs(), optional=True))


def build_day_block(
    report: dict[str, Any],
    *,
    previous_report: dict[str, Any] | None = None,
    start_row: int = 1,
    start_col: int = 1,
) -> list[list[Any]]:
    economics = report["economics"]
    note = report["note"]
    stock = report["stock"]
    metrics = report["metrics"]
    promo_status_value = (note.promo_status or "").strip() or "Не участвуем"
    negative_feedback_value = (note.negative_feedback or "").strip() or "Без изменений"
    ads_enabled_value = _bool(
        bool(note.unified_enabled or note.manual_search_enabled or note.manual_shelves_enabled)
    )
    price_change_status_value = (
        (getattr(note, "price_change_status", "") or "").strip()
        or ("Повысили" if note.price_changed else "Нет")
    )

    blocks = report["blocks"]
    total_ad = report["total_ad"]
    unified_search = blocks["unified_search"]
    unified_shelves = blocks["unified_shelves"]
    unified_catalog = blocks["unified_catalog"]
    manual_search = blocks["manual_search"]
    manual_catalog = blocks["manual_catalog"]
    manual_shelves = blocks["manual_shelves"]

    columns = [
        unified_search,
        unified_shelves,
        unified_catalog,
        manual_search,
        manual_catalog,
        manual_shelves,
    ]

    def has_visible_zone_traffic(cell: MetricCell) -> bool:
        return (
            has_metric_cell_data(cell)
            and (
                cell.impressions > 0
                or cell.clicks > 0
                or decimalize(cell.spend) != 0
            )
        )

    unified_group_visible = any(has_metric_cell_data(cell) for cell in [unified_search, unified_shelves, unified_catalog])
    ad_group_visible = any(has_metric_cell_data(cell) for cell in columns)
    active_columns = [
        unified_group_visible and has_visible_zone_traffic(unified_search),
        unified_group_visible and has_visible_zone_traffic(unified_shelves),
        unified_group_visible and has_visible_zone_traffic(unified_catalog),
        has_visible_zone_traffic(manual_search),
        has_visible_zone_traffic(manual_catalog),
        has_visible_zone_traffic(manual_shelves),
    ]

    unified_spend = (
        decimalize(unified_search.spend)
        + decimalize(unified_shelves.spend)
        + decimalize(unified_catalog.spend)
    )
    unified_impressions = unified_search.impressions + unified_shelves.impressions + unified_catalog.impressions
    unified_clicks = unified_search.clicks + unified_shelves.clicks + unified_catalog.clicks
    unified_carts = total_ad.carts  # Используем сумму всех рекламных кампаний
    unified_orders = total_ad.orders
    unified_order_sum = decimalize(total_ad.order_sum)
    
    # Органические (чистые) метрики = всего - реклама
    organic_clicks = max((metrics.open_count if metrics else 0) - total_ad.clicks, 0)
    organic_carts = max((metrics.add_to_cart_count if metrics else 0) - total_ad.carts, 0)
    organic_orders = max((metrics.order_count if metrics else 0) - total_ad.orders, 0)
    organic_order_sum = max(decimalize(metrics.order_sum if metrics else 0) - unified_order_sum, Decimal("0.00"))

    def pick_numbers(metric_name: str) -> list[int | float | str]:
        values: list[int | float | str] = []
        for index, cell in enumerate(columns):
            # Проверяем и группу, и конкретную ячейку (чтобы пустые колонки не показывались)
            if not active_columns[index] or not has_metric_cell_data(cell):
                values.append("")
                continue
            value = getattr(cell, metric_name)
            if isinstance(value, Decimal):
                values.append(_money(value, optional=True))
            else:
                values.append(_int(value, optional=True))
        return values

    def unified_traffic_value(relative_col: int) -> float | str:
        cell = columns[relative_col - 2]
        if not active_columns[relative_col - 2] or unified_impressions <= 0 or cell.impressions <= 0:
            return ""
        return _fraction(decimalize(cell.impressions) * Decimal("100") / decimalize(unified_impressions))

    def manual_traffic_value(relative_col: int) -> float | str:
        cell = columns[relative_col - 2]
        manual_total = manual_search.impressions + manual_catalog.impressions
        if not active_columns[relative_col - 2] or manual_total <= 0 or cell.impressions <= 0:
            return ""
        return _fraction(decimalize(cell.impressions) * Decimal("100") / decimalize(manual_total))

    def row_value_ref(relative_row: int, relative_col: int) -> str:
        return _cell_ref(
            start_row=start_row,
            start_col=start_col,
            relative_row=relative_row,
            relative_col=relative_col,
        )

    buyout_percent_ref = row_value_ref(22, 2)

    def buyout_formula(relative_col: int) -> str:
        return f"={row_value_ref(15, relative_col)}*{buyout_percent_ref}"

    def cost_per_order_formula(relative_col: int) -> str:
        return safe_divide_formula(
            row_value_ref(5, relative_col),
            row_value_ref(13, relative_col),
            treat_zero_numerator_as_empty=True,
        )

    def cost_per_cart_formula(relative_col: int) -> str:
        return safe_divide_formula(
            row_value_ref(5, relative_col),
            row_value_ref(11, relative_col),
            treat_zero_numerator_as_empty=True,
        )

    def ratio_formula(relative_col: int, denominator_row: int) -> str:
        return safe_divide_formula(
            row_value_ref(5, relative_col),
            row_value_ref(denominator_row, relative_col),
            scale=100,
            treat_zero_numerator_as_empty=True,
        )

    def profit_formula() -> str:
        keyword_count = len(report.get("keyword_rows") or [])
        seller_price_ref = row_value_ref(39 + keyword_count, 6)
        unit_cost_ref = row_value_ref(23, 2)
        logistics_ref = row_value_ref(24, 2)
        drr_sales_ref = row_value_ref(20, 8)
        orders_ref = row_value_ref(13, 8)
        buyout_fraction = f"IF({buyout_percent_ref}>1,{buyout_percent_ref}/100,{buyout_percent_ref})"
        return (
            f'=IFERROR(IF({seller_price_ref}=0,0,'
            f'(({seller_price_ref}-{unit_cost_ref}-({seller_price_ref}*({drr_sales_ref}/100)))'
            f'-({seller_price_ref}*25/100)-(({logistics_ref}/{buyout_fraction})-50))'
            f'*({orders_ref}*{buyout_fraction})),0)'
        )

    def safe_divide_formula(
        numerator_ref: str,
        denominator_ref: str,
        *,
        scale: int | None = None,
        fallback: str = '"-"',
        treat_zero_numerator_as_empty: bool = False,
    ) -> str:
        expression = f"{numerator_ref}/{denominator_ref}"
        if scale is not None:
            expression = f"({expression})*{scale}"
        conditions = [f"{denominator_ref}=\"\"", f"{denominator_ref}=0"]
        if treat_zero_numerator_as_empty:
            conditions.append(f"{numerator_ref}=\"\"")
            conditions.append(f"{numerator_ref}=0")
        joined_conditions = ",".join(conditions)
        return f'=IFERROR(IF(OR({joined_conditions}),{fallback},{expression}),{fallback})'

    def maybe_formula(relative_col: int, formula: str) -> str:
        return formula if active_columns[relative_col - 2] else ""

    def average_orders_formula() -> float | str:
        if stock and decimalize(stock.avg_orders_per_day):
            return _money(stock.avg_orders_per_day)
        block_span = BLOCK_WIDTH + BLOCK_GAP
        refs: list[str] = []
        for offset in range(7):
            block_start_col = start_col - offset * block_span
            if block_start_col < 1:
                break
            refs.append(
                _cell_ref(
                    start_row=start_row,
                    start_col=block_start_col,
                    relative_row=13,
                    relative_col=8,
                )
            )
        if len(refs) <= 1:
            return _money(report["avg_orders_per_day"])
        return f"=(({'+'.join(refs)})/{len(refs)})"

    def average_stock_drop_value() -> float | str:
        return _money(report["avg_stock_drop_per_day"], optional=True)

    def days_until_zero_value() -> float | str:
        return _money(report["days_until_zero_from_stock_drop"], optional=True)

    spp_delta_label, spp_delta_value = _spp_delta_parts(report, previous_report)
    spp_delta_text = spp_delta_value or spp_delta_label

    rows: list[list[Any]] = [
        ["", _block_header(report), "", "", "", "", "", "", ""],
        ["Тип рекламной кампании", "Единая ставка", "", "", "РС Поиск", "", "РС Полки", "Общая", "Органика"],
        ["Зоны показов", "Поиск", "Полки", "Каталог", "Поиск", "Каталог", "Полки", "Общая", "Органика"],
        [
            "Доля трафика (%)",
            unified_traffic_value(2),
            unified_traffic_value(3),
            unified_traffic_value(4),
            manual_traffic_value(5),
            manual_traffic_value(6),
            "",
            1 if ad_group_visible else "",
            "-",
        ],
        ["Затраты (руб)", *pick_numbers("spend"), f"=SUM({row_value_ref(5, 2)}:{row_value_ref(5, 7)})", "-"],
        ["Показы", *pick_numbers("impressions"), f"=SUM({row_value_ref(6, 2)}:{row_value_ref(6, 7)})", "-"],
        [
            "CTR",
            maybe_formula(2, safe_divide_formula(row_value_ref(10, 2), row_value_ref(6, 2), scale=100)),
            maybe_formula(3, safe_divide_formula(row_value_ref(10, 3), row_value_ref(6, 3), scale=100)),
            maybe_formula(4, safe_divide_formula(row_value_ref(10, 4), row_value_ref(6, 4), scale=100)),
            maybe_formula(5, safe_divide_formula(row_value_ref(10, 5), row_value_ref(6, 5), scale=100)),
            maybe_formula(6, safe_divide_formula(row_value_ref(10, 6), row_value_ref(6, 6), scale=100)),
            maybe_formula(7, safe_divide_formula(row_value_ref(10, 7), row_value_ref(6, 7), scale=100)),
            safe_divide_formula(f"SUM({row_value_ref(10, 2)}:{row_value_ref(10, 7)})", row_value_ref(6, 8), scale=100),
            "-",
        ],
        [
            "CPM",
            maybe_formula(2, safe_divide_formula(f"{row_value_ref(5, 2)}*1000", row_value_ref(6, 2))),
            maybe_formula(3, safe_divide_formula(f"{row_value_ref(5, 3)}*1000", row_value_ref(6, 3))),
            maybe_formula(4, safe_divide_formula(f"{row_value_ref(5, 4)}*1000", row_value_ref(6, 4))),
            maybe_formula(5, safe_divide_formula(f"{row_value_ref(5, 5)}*1000", row_value_ref(6, 5))),
            maybe_formula(6, safe_divide_formula(f"{row_value_ref(5, 6)}*1000", row_value_ref(6, 6))),
            maybe_formula(7, safe_divide_formula(f"{row_value_ref(5, 7)}*1000", row_value_ref(6, 7))),
            safe_divide_formula(f"{row_value_ref(5, 8)}*1000", row_value_ref(6, 8)),
            "-",
        ],
        [
            "CPC",
            maybe_formula(2, safe_divide_formula(row_value_ref(5, 2), row_value_ref(10, 2))),
            maybe_formula(3, safe_divide_formula(row_value_ref(5, 3), row_value_ref(10, 3))),
            maybe_formula(4, safe_divide_formula(row_value_ref(5, 4), row_value_ref(10, 4))),
            maybe_formula(5, safe_divide_formula(row_value_ref(5, 5), row_value_ref(10, 5))),
            maybe_formula(6, safe_divide_formula(row_value_ref(5, 6), row_value_ref(10, 6))),
            maybe_formula(7, safe_divide_formula(row_value_ref(5, 7), row_value_ref(10, 7))),
            safe_divide_formula(row_value_ref(5, 8), f"SUM({row_value_ref(10, 2)}:{row_value_ref(10, 7)})"),
            "-",
        ],
        ["Клики", *pick_numbers("clicks"), _int(metrics.open_count if metrics else 0), f"={row_value_ref(10, 8)}-SUM({row_value_ref(10, 2)}:{row_value_ref(10, 7)})"],
        ["Корзины", *pick_numbers("carts"), _int(metrics.add_to_cart_count if metrics else 0), f"={row_value_ref(11, 8)}-SUM({row_value_ref(11, 2)}:{row_value_ref(11, 7)})"],
        ["Конверсия в корзину (%)", "", "", "", "", "", "", safe_divide_formula(row_value_ref(11, 8), row_value_ref(10, 8), scale=100), ""],
        ["Заказы", *pick_numbers("orders"), _int(metrics.order_count if metrics else 0), f"={row_value_ref(13, 8)}-SUM({row_value_ref(13, 2)}:{row_value_ref(13, 7)})"],
        ["Конверсия в заказ (%)", "", "", "", "", "", "", safe_divide_formula(row_value_ref(13, 8), row_value_ref(11, 8), scale=100), ""],
        ["Заказы (руб.)", *pick_numbers("order_sum"), _money(metrics.order_sum if metrics else 0), f"={row_value_ref(15, 8)}-SUM({row_value_ref(15, 2)}:{row_value_ref(15, 7)})"],
        ["Выкупы ≈ (руб.)", maybe_formula(2, buyout_formula(2)), maybe_formula(3, buyout_formula(3)), maybe_formula(4, buyout_formula(4)), maybe_formula(5, buyout_formula(5)), maybe_formula(6, buyout_formula(6)), maybe_formula(7, buyout_formula(7)), buyout_formula(8), "-"],
        ["Стоимость заказа", maybe_formula(2, cost_per_order_formula(2)), maybe_formula(3, cost_per_order_formula(3)), maybe_formula(4, cost_per_order_formula(4)), maybe_formula(5, cost_per_order_formula(5)), maybe_formula(6, cost_per_order_formula(6)), maybe_formula(7, cost_per_order_formula(7)), safe_divide_formula(row_value_ref(5, 8), row_value_ref(13, 8), treat_zero_numerator_as_empty=True), "-"],
        ["Стоимость корзины", maybe_formula(2, cost_per_cart_formula(2)), maybe_formula(3, cost_per_cart_formula(3)), maybe_formula(4, cost_per_cart_formula(4)), maybe_formula(5, cost_per_cart_formula(5)), maybe_formula(6, cost_per_cart_formula(6)), maybe_formula(7, cost_per_cart_formula(7)), safe_divide_formula(row_value_ref(5, 8), row_value_ref(11, 8), treat_zero_numerator_as_empty=True), "-"],
        ["ДРР от заказов (%)", maybe_formula(2, ratio_formula(2, 15)), maybe_formula(3, ratio_formula(3, 15)), maybe_formula(4, ratio_formula(4, 15)), maybe_formula(5, ratio_formula(5, 15)), maybe_formula(6, ratio_formula(6, 15)), maybe_formula(7, ratio_formula(7, 15)), safe_divide_formula(row_value_ref(5, 8), row_value_ref(15, 8), scale=100, treat_zero_numerator_as_empty=True), "-"],
        ["ДРР от продаж ≈ (%)", maybe_formula(2, ratio_formula(2, 16)), maybe_formula(3, ratio_formula(3, 16)), maybe_formula(4, ratio_formula(4, 16)), maybe_formula(5, ratio_formula(5, 16)), maybe_formula(6, ratio_formula(6, 16)), maybe_formula(7, ratio_formula(7, 16)), safe_divide_formula(row_value_ref(5, 8), row_value_ref(16, 8), scale=100, treat_zero_numerator_as_empty=True), "-"],
        ["прибыль (без налогов, костов вне ВБ и возвратов)", profit_formula(), "", "", "", "", "", "", ""],
        ["Процент выкупа %", _fraction(economics.buyout_percent), "", "", "", "", "", "", ""],
        ["Себестоимость", _money(economics.unit_cost, optional=True), "", "", "", "", "", "", ""],
        ["Логистика", _money(economics.logistics_cost, optional=True), "", "", "", "", "", "", ""],
        ["Остатки:", "", "", "", "", "", "", "", ""],
        ["Остатки на складах WB", "", "", "", _int(stock.total_stock if stock else 0), "", "", "", ""],
        ["Едут к клиенту", "", "", "", _int(stock.in_way_to_client if stock else 0), "", "", "", ""],
        ["Возвращаются на склад", "", "", "", _int(stock.in_way_from_client if stock else 0), "", "", "", ""],
        ["Ср. кол-во заказов/день", "", "", "", average_orders_formula(), "", "", "", ""],
        ["Ср. убыль остатков/день", "", "", "", average_stock_drop_value(), "", "", "", ""],
        ["Дней до АУТА", "", "", "", days_until_zero_value(), "", "", "", ""],
        ["Остатки по складам", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
    ]
    rows.append(["Ключи", "Частота", "Позиция ОРГАНИЧЕСКАЯ", "", "Позиция БУСТОВАЯ", "", "", "CTR (%)", ""])
    for keyword_row in report.get("keyword_rows") or []:
        has_data = bool(keyword_row.get("has_data"))
        rows.append(
            [
                keyword_row.get("query_text") or "",
                _int(keyword_row.get("frequency"), optional=not has_data),
                _money(keyword_row.get("organic_position"), optional=not has_data),
                "",
                _money(keyword_row.get("boosted_position"), optional=not has_data),
                "",
                "",
                _fraction(keyword_row.get("boosted_ctr"), optional=not has_data),
                "",
            ]
        )
    rows.extend(
        [
            ["Обзор:", "", "", "", "", "", "", "", ""],
            ["СПП", "", "", _fraction(note.spp_percent, optional=True), "", spp_delta_label, "", spp_delta_value, ""],
            ["Цена WBSELLER (наша)", "", "", "", "", _money(note.seller_price, optional=True), "", "", ""],
            ["Цена WB (на сайте)", "", "", "", "", _money(note.wb_price, optional=True), "", "", ""],
            ["Акция", "", "", "", "", promo_status_value, "", "", ""],
            ["Негативные отзывы", "", "", "", "", negative_feedback_value, "", "", ""],
            ["Действия:", "", "", "", "", "", "", "", ""],
            ["Включили рекламу?", "", "", "", ads_enabled_value, "", "", _money(getattr(note, "ads_budget", 0), optional=True), ""],
            ["Меняли цену?(WBSeller)", "", "", "", price_change_status_value, "", "", _money(getattr(note, "price_change_amount", 0), optional=True), ""],
            ["Комментарии:", "", "", "", "", "", "", "", ""],
            [note.comment, "", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", "", ""],
        ]
    )
    return rows


def _build_prefetched_product_report_context(*, product: Product, stock_dates: list[date]) -> dict[str, Any]:
    if not stock_dates:
        return {}

    normalized_dates = sorted(set(stock_dates))
    max_stock_date = normalized_dates[-1]

    metrics_by_date = {
        row.stats_date: row
        for row in DailyProductMetrics.objects.filter(
            product=product,
            stats_date__in=normalized_dates,
        )
    }
    stocks_by_date = {
        row.stats_date: row
        for row in DailyProductStock.objects.filter(
            product=product,
            stats_date__in=normalized_dates,
        )
    }
    notes_by_date = {
        row.note_date: row
        for row in DailyProductNote.objects.filter(
            product=product,
            note_date__in=normalized_dates,
        )
    }

    warehouse_rows_by_date: dict[date, list[DailyWarehouseStock]] = defaultdict(list)
    for row in DailyWarehouseStock.objects.filter(
        product=product,
        stats_date__in=normalized_dates,
    ).select_related("warehouse"):
        warehouse_rows_by_date[row.stats_date].append(row)

    campaign_stats_by_date: dict[date, list[DailyCampaignProductStat]] = defaultdict(list)
    for row in (
        DailyCampaignProductStat.objects.filter(
            product=product,
            stats_date__in=normalized_dates,
        )
        .select_related("campaign")
        .order_by("stats_date", "campaign__monitoring_group", "zone")
    ):
        campaign_stats_by_date[row.stats_date].append(row)

    keyword_stats_by_date: dict[date, list[DailyProductKeywordStat]] = defaultdict(list)
    for row in DailyProductKeywordStat.objects.filter(
        product=product,
        stats_date__in=normalized_dates,
    ).order_by("stats_date", "query_text"):
        keyword_stats_by_date[row.stats_date].append(row)
    product_keywords = list(
        ProductKeyword.objects.filter(product=product)
        .order_by("position", "query_text", "id")
        .values_list("query_text", flat=True)
    )

    economics_versions = list(
        ProductEconomicsVersion.objects.filter(
            product=product,
            effective_from__lte=max_stock_date,
        ).order_by("-effective_from", "-id")
    )
    fallback_buyout = decimalize(product.buyout_percent)
    fallback_unit_cost = decimalize(product.unit_cost)
    fallback_logistics = decimalize(product.logistics_cost)
    economics_by_date: dict[date, ResolvedEconomics] = {}
    for current_date in normalized_dates:
        economics_snapshot = next(
            (version for version in economics_versions if version.effective_from <= current_date),
            None,
        )
        if economics_snapshot:
            resolved = ResolvedEconomics(
                effective_from=economics_snapshot.effective_from,
                buyout_percent=decimalize(economics_snapshot.buyout_percent),
                unit_cost=decimalize(economics_snapshot.unit_cost),
                logistics_cost=decimalize(economics_snapshot.logistics_cost),
            )
        else:
            resolved = ResolvedEconomics(
                effective_from=None,
                buyout_percent=decimalize(product.buyout_percent),
                unit_cost=decimalize(product.unit_cost),
                logistics_cost=decimalize(product.logistics_cost),
            )
        economics_by_date[current_date] = ResolvedEconomics(
            effective_from=resolved.effective_from,
            buyout_percent=fallback_buyout if resolved.buyout_percent == 0 and fallback_buyout != 0 else resolved.buyout_percent,
            unit_cost=fallback_unit_cost if resolved.unit_cost == 0 and fallback_unit_cost != 0 else resolved.unit_cost,
            logistics_cost=fallback_logistics if resolved.logistics_cost == 0 and fallback_logistics != 0 else resolved.logistics_cost,
        )

    visible_warehouse_names = {
        normalize_warehouse_name(name)
        for name in product.visible_warehouse_names()
    }
    active_campaign_exists = product.campaigns.filter(is_active=True).exists()

    history_rows = list(product.daily_metrics.order_by("-stats_date")[:14])
    rolling_avg_orders = Decimal("0")
    if history_rows:
        window = history_rows[:7]
        rolling_avg_orders = sum((Decimal(item.order_count) for item in window), Decimal("0")) / Decimal(len(window))

    return {
        "preloaded_metrics": metrics_by_date,
        "preloaded_stocks": stocks_by_date,
        "preloaded_notes": notes_by_date,
        "preloaded_warehouse_rows": warehouse_rows_by_date,
        "preloaded_campaign_stats": campaign_stats_by_date,
        "preloaded_keyword_stats": keyword_stats_by_date,
        "preloaded_product_keywords": product_keywords,
        "preloaded_economics": economics_by_date,
        "preloaded_visible_warehouse_names": visible_warehouse_names,
        "preloaded_active_campaign_exists": active_campaign_exists,
        "preloaded_history": history_rows,
        "preloaded_rolling_avg_orders": rolling_avg_orders,
    }


def build_product_monitoring_rows(*, product: Product, reference_date: date, history_days: int) -> list[list[Any]]:
    stock_dates = [reference_date - timedelta(days=offset) for offset in reversed(range(history_days))]
    prefetched_context = _build_prefetched_product_report_context(product=product, stock_dates=stock_dates)
    reports = [
        build_product_report(
            product=product,
            stats_date=stock_date,
            stock_date=stock_date,
            create_note=False,
            **prefetched_context,
        )
        for stock_date in stock_dates
    ]
    _normalize_keyword_rows_count(reports)

    block_height = _block_height(reports[0]["keyword_rows"]) if reports else BASE_BLOCK_HEIGHT
    total_width = history_days * BLOCK_WIDTH + max(history_days - 1, 0) * BLOCK_GAP
    matrix = [["" for _ in range(total_width)] for _ in range(block_height)]

    for index, report in enumerate(reports):
        start_col = 1 + index * (BLOCK_WIDTH + BLOCK_GAP)
        block = build_day_block(
            report,
            previous_report=reports[index - 1] if index > 0 else None,
            start_row=1,
            start_col=start_col,
        )
        for row_offset, block_row in enumerate(block):
            if row_offset >= block_height:
                break
            for col_offset, value in enumerate(block_row):
                keep_repeated_stock_label = 24 <= row_offset <= 31
                if index > 0 and col_offset == 0 and not keep_repeated_stock_label:
                    value = ""
                matrix[row_offset][start_col + col_offset - 1] = value
    return matrix


def build_product_monitoring_rows_display(
    *,
    product: Product,
    reference_date: date,
    history_days: int,
    stock_dates: list[date] | None = None,
) -> list[list[Any]]:
    resolved_stock_dates = stock_dates or [reference_date - timedelta(days=offset) for offset in reversed(range(history_days))]
    resolved_history_days = len(resolved_stock_dates)
    prefetched_context = _build_prefetched_product_report_context(product=product, stock_dates=resolved_stock_dates)
    reports = [
        build_product_report(
            product=product,
            stats_date=stock_date,
            stock_date=stock_date,
            create_note=False,
            **prefetched_context,
        )
        for stock_date in resolved_stock_dates
    ]
    _normalize_keyword_rows_count(reports)

    block_height = _block_height(reports[0]["keyword_rows"]) if reports else BASE_BLOCK_HEIGHT
    keyword_count = len(reports[0]["keyword_rows"]) if reports else 0
    footer_start_offset = 36 + keyword_count
    total_width = resolved_history_days * BLOCK_WIDTH + max(resolved_history_days - 1, 0) * BLOCK_GAP
    matrix = [["" for _ in range(total_width)] for _ in range(block_height)]

    for index, report in enumerate(reports):
        previous_report = reports[index - 1] if index > 0 else None
        start_col = index * (BLOCK_WIDTH + BLOCK_GAP)
        block = exporter_rows(report, previous_report=previous_report)
        for row_offset, block_row in enumerate(block):
            if row_offset >= block_height:
                break
            for col_offset, value in enumerate(block_row):
                keep_repeated_stock_label = 24 <= row_offset <= 31
                keep_repeated_footer_label = row_offset >= footer_start_offset
                if index > 0 and col_offset == 0 and not (keep_repeated_stock_label or keep_repeated_footer_label):
                    value = ""
                target_col = start_col + col_offset
                if target_col >= total_width:
                    break
                matrix[row_offset][target_col] = value

    return matrix


def build_dashboard_rows(*, reference_date: date, history_days: int, product_ids: list[int] | None = None) -> list[list[Any]]:
    stats_date = reference_date
    context = build_dashboard_context(stats_date=stats_date, stock_date=reference_date)
    settings = get_monitoring_settings()
    rows: list[list[Any]] = [
        [settings.project_name, "", "", "", "", ""],
        ["Сформировано", timezone.localtime().strftime("%Y-%m-%d %H:%M"), "", "", "", ""],
        ["Глубина по дням", reference_date.isoformat(), "Окно, дней", history_days, "", ""],
        ["Дата рекламной статистики", stats_date.isoformat(), "", "", "", ""],
        [],
        ["Лист", "nmID", "Товар", "Артикул продавца", "Заказы", "Расход РК", "Остаток WB", "К клиенту", "Кампаний"],
    ]
    for card in context["cards"]:
        if product_ids and card["product"].id not in product_ids:
            continue
        rows.append(
            [
                monitoring_sheet_title(card["product"]),
                card["product"].nm_id,
                card["product"].title,
                card["product"].vendor_code,
                card["metrics"].order_count if card["metrics"] else 0,
                _money(card["total_spend"]),
                card["stock"].total_stock if card["stock"] else 0,
                card["stock"].in_way_to_client if card["stock"] else 0,
                card["campaigns_count"],
            ]
        )
    return rows


def _products_for_export(product_ids: list[int] | None = None) -> list[Product]:
    queryset = Product.objects.filter(is_active=True)
    if product_ids:
        queryset = queryset.filter(id__in=product_ids)
    return list(queryset.order_by("vendor_code", "title", "nm_id"))


def build_monitoring_sheet_payloads(
    *,
    reference_date: date,
    history_days: int | None = None,
    product_ids: list[int] | None = None,
) -> list[MonitoringSheetPayload]:
    settings = get_monitoring_settings()
    resolved_history_days = history_days or getattr(settings, "monitoring_history_days", 14) or 14
    products = _products_for_export(product_ids)
    payloads = [
        MonitoringSheetPayload(
            title=normalize_title("Dashboard"),
            rows=build_dashboard_rows(reference_date=reference_date, history_days=resolved_history_days, product_ids=product_ids),
            kind="dashboard",
        )
    ]
    for product in products:
        payloads.append(
            MonitoringSheetPayload(
                title=monitoring_sheet_title(product),
                rows=build_product_monitoring_rows(
                    product=product,
                    reference_date=reference_date,
                    history_days=resolved_history_days,
                ),
                kind="product",
            )
        )
    return payloads


def build_table_view_payloads(
    *,
    reference_date: date,
    history_days: int | None = None,
    product_ids: list[int] | None = None,
    active_sheet_key: str | None = None,
) -> list[MonitoringSheetPayload]:
    settings = get_monitoring_settings()
    resolved_history_days = history_days or getattr(settings, "monitoring_history_days", 14) or 14
    products = _products_for_export(product_ids)
    active_key = (active_sheet_key or "").strip()

    def is_active(sheet_index: int) -> bool:
        if not active_key:
            return True
        return active_key == f"sheet-{sheet_index}"

    payloads = [
        MonitoringSheetPayload(
            title=normalize_title("Dashboard"),
            rows=(
                build_dashboard_rows(
                    reference_date=reference_date,
                    history_days=resolved_history_days,
                    product_ids=product_ids,
                )
                if is_active(0)
                else []
            ),
            kind="dashboard",
        )
    ]
    stock_dates = [reference_date - timedelta(days=offset) for offset in reversed(range(resolved_history_days))]
    for product in products:
        sheet_index = len(payloads)
        payloads.append(
            MonitoringSheetPayload(
                title=monitoring_sheet_title(product),
                rows=(
                    build_product_monitoring_rows_display(
                        product=product,
                        reference_date=reference_date,
                        history_days=resolved_history_days,
                        stock_dates=stock_dates,
                    )
                    if is_active(sheet_index)
                    else []
                ),
                kind="product",
                product_id=product.id,
                block_dates=stock_dates,
            )
        )
    return payloads


def _apply_dashboard_style(sheet) -> None:
    sheet.freeze_panes = "A6"
    widths = {1: 24, 2: 14, 3: 32, 4: 18, 5: 12, 6: 14, 7: 12, 8: 12, 9: 10}
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif cell.row in (2, 3, 4, 6):
                cell.fill = SUBHEADER_FILL
                cell.font = SECTION_FONT if cell.row == 6 else Font(bold=True)


def _apply_product_sheet_style(sheet, history_days: int, block_height: int) -> None:
    sheet.freeze_panes = "B1"
    keyword_offset = max(0, block_height - BASE_BLOCK_HEIGHT)
    keyword_header_row = 36
    overview_row = 37 + keyword_offset
    actions_row = overview_row + 6
    comments_row = overview_row + 9
    percent_rows = {4, 12, 14, 19, 20, 22, overview_row + 1}
    money_rows = {
        5,
        15,
        16,
        17,
        18,
        21,
        23,
        24,
        overview_row + 2,
        overview_row + 3,
        actions_row + 1,
        actions_row + 2,
    }
    integer_rows = {6, 10, 11, 13, 26, 27, 28}
    decimal_rows = {7, 8, 9, 29, 30, 31}
    muted_rows = {5, 7, 9, 11, 13, 15, 17, 19, 21, 23}
    dark_rows = {1, 2, 25, keyword_header_row, overview_row, actions_row, comments_row}
    screenshot_rows = set(range(32, 36))

    for row_idx in range(1, block_height + 1):
        if row_idx in (1, 2, 3):
            sheet.row_dimensions[row_idx].height = 24
        elif row_idx in dark_rows:
            sheet.row_dimensions[row_idx].height = 22
        elif row_idx in screenshot_rows:
            sheet.row_dimensions[row_idx].height = 26
        else:
            sheet.row_dimensions[row_idx].height = 19

    for block_index in range(history_days):
        start_col = 1 + block_index * (BLOCK_WIDTH + BLOCK_GAP)
        widths = [27.63 if block_index == 0 else 5.38, 11.38, 13, 13, 13, 13, 13, 13, 13]
        for offset, width in enumerate(widths):
            sheet.column_dimensions[get_column_letter(start_col + offset)].width = width
        if BLOCK_GAP and block_index < history_days - 1:
            sheet.column_dimensions[get_column_letter(start_col + BLOCK_WIDTH)].width = 4

        for row_idx in range(1, block_height + 1):
            for col_offset in range(BLOCK_WIDTH):
                cell = sheet.cell(row=row_idx, column=start_col + col_offset)
                if col_offset == 0:
                    cell.border = BLOCK_LEFT_BORDER
                elif col_offset == BLOCK_WIDTH - 1:
                    cell.border = BLOCK_RIGHT_BORDER
                else:
                    cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

                if row_idx in dark_rows:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                elif row_idx == 3 or row_idx in muted_rows or row_idx == 26:
                    cell.fill = SUBHEADER_FILL
                    cell.font = Font(bold=col_offset == 0 or row_idx in {3, 21, 23, 26})
                elif row_idx in screenshot_rows:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font(name="Arial")
                else:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font(name="Arial", bold=col_offset == 0)

                if row_idx in percent_rows:
                    cell.number_format = "0.##%"
                elif row_idx in money_rows:
                    cell.number_format = "#,##0.##"
                elif row_idx in integer_rows:
                    cell.number_format = "#,##0"
                elif row_idx in decimal_rows:
                    cell.number_format = "#,##0.##"

                if row_idx == 21:
                    cell.font = Font(name="Arial", size=14, bold=True, color="34A853")
                elif row_idx in {5}:
                    cell.font = Font(name="Arial", bold=col_offset == 0, color="34A853")
                elif row_idx in {11, 12}:
                    cell.font = Font(name="Arial", bold=col_offset == 0 or row_idx == 12, color="4285F4")
                elif row_idx in {13, 14}:
                    cell.font = Font(name="Arial", bold=col_offset == 0 or row_idx == 14, color="FF9900")

                if col_offset == 0 or row_idx in dark_rows:
                    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

        def merge(row_idx: int, first_offset: int, last_offset: int, end_row: int | None = None) -> None:
            sheet.merge_cells(
                start_row=row_idx,
                start_column=start_col + first_offset,
                end_row=end_row or row_idx,
                end_column=start_col + last_offset,
            )

        merge(1, 1, 8)
        merge(2, 1, 3)
        merge(2, 4, 5)
        for row_idx in range(21, 25):
            merge(row_idx, 1, 8)
        merge(25, 0, 8)
        for row_idx in range(26, 32):
            merge(row_idx, 0, 3)
            merge(row_idx, 4, 8)
        merge(32, 0, 8, 35)
        for row_idx in range(keyword_header_row, overview_row):
            merge(row_idx, 2, 3)
            merge(row_idx, 4, 5)
            merge(row_idx, 7, 8)
        merge(overview_row, 0, 8)
        merge(overview_row + 1, 0, 2)
        merge(overview_row + 1, 3, 4)
        merge(overview_row + 1, 5, 6)
        merge(overview_row + 1, 7, 8)
        for row_idx in range(overview_row + 2, overview_row + 6):
            merge(row_idx, 0, 4)
            merge(row_idx, 5, 8)
        merge(actions_row, 0, 8)
        for row_idx in range(actions_row + 1, actions_row + 3):
            merge(row_idx, 0, 3)
            merge(row_idx, 4, 5)
            merge(row_idx, 7, 8)
        merge(comments_row, 0, 8)
        merge(comments_row + 1, 0, 8, comments_row + 4)

        validation_specs = [
            (overview_row + 1, 5, '"Без изменений,Вырос на,Упал на"'),
            (overview_row + 4, 5, '"Участвуем,Не участвуем"'),
            (overview_row + 5, 5, '"Без изменений,Поступили"'),
            (actions_row + 1, 4, '"Да,Нет"'),
            (actions_row + 2, 4, '"Нет,Повысили,Понизили"'),
        ]
        for row_idx, col_offset, formula in validation_specs:
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(sheet.cell(row=row_idx, column=start_col + col_offset))

        if BLOCK_GAP and block_index < history_days - 1:
            separator_col = start_col + BLOCK_WIDTH
            for row_idx in range(1, block_height + 1):
                separator_cell = sheet.cell(row=row_idx, column=separator_col)
                separator_cell.value = ""
                separator_cell.fill = BLOCK_SEPARATOR_FILL
                separator_cell.border = NO_BORDER
    for row_idx in range(comments_row + 1, comments_row + 5):
        sheet.row_dimensions[row_idx].height = 14.25


def build_monitoring_workbook(
    *,
    reference_date: date,
    history_days: int | None = None,
    product_ids: list[int] | None = None,
) -> Workbook:
    payloads = build_monitoring_sheet_payloads(
        reference_date=reference_date,
        history_days=history_days,
        product_ids=product_ids,
    )
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    resolved_history_days = history_days or getattr(get_monitoring_settings(), "monitoring_history_days", 14) or 14
    for payload in payloads:
        sheet = workbook.create_sheet(title=payload.title)
        for row_idx, row in enumerate(payload.rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        if payload.kind == "dashboard":
            _apply_dashboard_style(sheet)
        else:
            _apply_product_sheet_style(sheet, resolved_history_days, sheet.max_row)
    return workbook


def export_monitoring_workbook_bytes(
    *,
    reference_date: date,
    history_days: int | None = None,
    product_ids: list[int] | None = None,
) -> bytes:
    workbook = build_monitoring_workbook(
        reference_date=reference_date,
        history_days=history_days,
        product_ids=product_ids,
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
